""" print worksheets at a time now! """
import sys
from collections import namedtuple
from copy import copy
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, Font, Border, PatternFill, Side, Alignment

from addresses import get_gather_address
from names import pretty_print_name
from print_territory import get_territory_docs, key_residence, phone_number_sort

def add_default_styles(workbook):
    """ plug in reusable styles """
    border = Side(style="thin", color="000000")
    thick_border = Side(style="thick", color="000000")

    default_style = NamedStyle(name="default_style")
    default_style.font = Font(size=15)
    default_style.border = Border(top=border, left=border, bottom=border, right=border)
    workbook.add_named_style(default_style)

    top_row = copy(default_style)
    top_row.name = "top_row"
    top_row.border = Border(top=thick_border, left=border, bottom=border, right=border)
    workbook.add_named_style(top_row)

    centered = Alignment(horizontal="center", vertical="center")
    for style in (default_style, top_row):
        style = copy(style)
        style.name = style.name + "_centered"
        style.alignment = centered
        workbook.add_named_style(style)

        style = copy(style)
        style.name = style.name + "_no_letter"
        style.fill = PatternFill("solid", fgColor="b7b7b7")
        workbook.add_named_style(style)


def create_workbook():
    """ create a workbook handle """
    return Workbook()


def open_workbook(path):
    """ open existing workbook """
    return load_workbook(filename=path)


def make_street_sheet(workbook, t_id, street):
    """ copies template and makes sure that we print everything """
    current_street_sheet = workbook.copy_worksheet(workbook["Street Name"])
    current_street_sheet.title = street
    write_to_sheet(current_street_sheet, 0, 1, t_id, None)
    return current_street_sheet


COLS = [chr(65 + i) for i in range(26)]

def write_to_sheet(sheet, col, row, val, style="default_style"):
    """ write to sheet """
    loc = "{}{}".format(COLS[col], row)
    sheet[loc] = val
    if style:
        sheet[loc].style = style
    return sheet[loc]


def write_dnc_to_sheet(sheet, col, row, val, style="default_style"):
    """ write dnc to sheet """
    cell = write_to_sheet(sheet, col, row, val, style)
    cell.fill = PatternFill("solid", fgColor="f4cccc")


def write_dnc_row(sheet, row, data):
    """ print a whole row at a time """
    style = "top_row" if row == 3 else "default_style"
    centered = style + "_centered"

    write_dnc_to_sheet(sheet, 0, row, data.name, style)
    write_dnc_to_sheet(sheet, 1, row, data.address, style)
    write_dnc_to_sheet(sheet, 2, row, "Do Not Call", style)
    write_dnc_to_sheet(sheet, 3, row, "DNC", centered)
    write_dnc_to_sheet(sheet, 4, row, "DNC", centered)
    write_dnc_to_sheet(sheet, 5, row, "DNC", centered)
    write_dnc_to_sheet(sheet, 6, row, "", style)
    return row + 1


def write_row(sheet, row, data):
    """ print a whole row at a time """
    style = "top_row" if row == 3 else "default_style"
    centered = style + "_centered"

    write_to_sheet(sheet, 0, row, data.name, style)
    write_to_sheet(sheet, 1, row, data.address, style)
    write_to_sheet(sheet, 2, row, data.phone, style)
    write_to_sheet(sheet, 3, row, data.call1, centered)
    write_to_sheet(sheet, 4, row, data.call2, centered)
    if data.letter == "⃠":
        write_to_sheet(sheet, 5, row, data.letter, centered + "_no_letter")
    else:
        write_to_sheet(sheet, 5, row, data.letter, centered)
    write_to_sheet(sheet, 6, row, data.note, style)
    return row + 1


ROW_FIELDS = ("name", "address", "phone", "call1", "call2", "letter", "note")
Row = namedtuple("Row", ROW_FIELDS, defaults=("",) * len(ROW_FIELDS))

def stamp_last_updated(workbook, newest_date):
    """ stamps the last updated date on the readme page """
    workbook["Read Me"]["C1"] = "Last Updated: {}".format(newest_date)


def remove_template_sheets(workbook):
    """ cleans up template sheets """
    for name in ("Street Name",):
        workbook.remove(workbook[name])

def get_phones_for_resident(resident):
    """ clean way to get unique, sorted phone numbers """
    phone_numbers = resident.get("phoneNumbers", [])

    if phone_numbers:
        return sorted(
            list({p["number"] for p in phone_numbers}),
            key=phone_number_sort
        )[0:2]

    return phone_numbers

def print_workbook(t_id):
    """  ready! to print """
    workbook = open_workbook("template.xlsx")
    add_default_styles(workbook)

    for street, residences in get_territory_docs(t_id):
        residences.sort(key=key_residence)
        row = 3
        sheet = make_street_sheet(workbook, t_id, street)
        youngest_change = datetime(2000, 1, 1)

        for resident in residences:
            addr = get_gather_address(resident["address"])
            name = pretty_print_name(resident.get("name", None) or "Current Resident")

            if youngest_change < resident.get("lastUpdate"):
                youngest_change = resident.get("lastUpdate")

            # do not call check
            if resident.get("doNotCall", None):
                row = write_dnc_row(sheet, row, Row(name, addr))
                continue

            phones = get_phones_for_resident(resident)
            if phones:
                for i, phone in enumerate(phones):
                    if i > 0:
                        row_data = Row("▲", "▲", phone, None, None, "⃠")
                    else:
                        row_data = Row(name, addr, phone)
                    row = write_row(sheet, row, row_data)

            else:
                row = write_row(sheet, row, Row(name, addr, "No Number Found", "-", "-"))

    stamp_last_updated(workbook, youngest_change)
    remove_template_sheets(workbook)
    workbook.save(filename="{}.xlsx".format(t_id))


if __name__ == "__main__" and len(sys.argv) > 1:
    print_workbook(sys.argv[1])
