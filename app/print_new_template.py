""" print worksheets at a time now! """
import sys
from collections import namedtuple
from copy import copy
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, Font, Border, PatternFill, Side, Alignment

from mongo import get_all_docs_for
from addresses import get_gather_address, get_street, get_sheet_street
from print_territory import key_residence

def get_territory_docs(territory_id):
    """ collects all docs for territory """
    territory = {}
    for doc in get_all_docs_for(territory_id):
        street = get_street(doc["address"])
        if street not in territory:
            territory[street] = []
        territory[street].append(doc)
    return territory.items()


def add_default_styles(workbook):
    """ plug in reusable styles """
    border = Side(style="thin", color="000000")
    thick_border = Side(style="thick", color="000000")

    default_style = NamedStyle(name="default_style")
    default_style.font = Font(name="Arial", size=10)
    default_style.border = Border(top=border, left=border, bottom=border, right=border)
    workbook.add_named_style(default_style)

    centered = Alignment(horizontal="center", vertical="center")

    street_style = NamedStyle(name="street")
    street_style.font = Font(name="Arial", size=10, bold=True)
    street_style.alignment = centered
    workbook.add_named_style(street_style)

    top_row = copy(default_style)
    top_row.name = "top_row"
    top_row.border = Border(top=thick_border, left=border, bottom=border, right=border)
    workbook.add_named_style(top_row)

    thick_right = copy(default_style)
    thick_right.name = "thick_right"
    thick_right.border = Border(top=border, left=border, bottom=border, right=thick_border)
    workbook.add_named_style(thick_right)

    for style in (default_style, top_row):
        style = copy(style)
        style.name = style.name + "_centered"
        style.alignment = centered
        workbook.add_named_style(style)

        style = copy(style)
        style.name = style.name + "_no_letter"
        style.fill = PatternFill("solid", fgColor="b7b7b7")
        workbook.add_named_style(style)


def create_workbook():
    """ create a workbook handle """
    return Workbook()


def open_workbook(path):
    """ open existing workbook """
    return load_workbook(filename=path)


def make_street_sheet(workbook, t_id, street, residences):
    """ copies template and makes sure that we print everything """
    current_street_sheet = workbook.copy_worksheet(workbook["Sheet1"])
    current_street_sheet.title = street
    write_to_sheet(current_street_sheet, 0, 1, t_id, None)
    current_street_sheet.row_dimensions[1].height = 72
    write_to_sheet(current_street_sheet, 1, 1, get_sheet_street(residences[0].get("address")), "street")
    return current_street_sheet


COLS = [chr(65 + i) for i in range(26)]

def write_to_sheet(sheet, col, row, val, style="default_style"):
    """ write to sheet """
    loc = "{}{}".format(COLS[col], row)
    if style and col == 15:
        sheet[loc].hyperlink = val[0]
        sheet[loc].value = val[1]
    else:
        sheet[loc] = val
    if style:
        sheet[loc].style = style
    return sheet[loc]


def write_row(sheet, row, data):
    """ print a whole row at a time """
    style = "default_style"
    centered = style + "_centered"
    if row > 5:
        sheet.insert_rows(row)
    conflict_message = ""
    if data.conflict:
        map_link = "http://www.google.com/maps/place/{}".format(",".join(data.coords))

        conflict_message = "Assessor address conflicts! Please verify correct address and update!"
        conflict_message = "{}\n{}".format(conflict_message, map_link)

    write_to_sheet(sheet, 0, row, "", style) #A
    write_to_sheet(sheet, 1, row, data.address, style) #B
    write_to_sheet(sheet, 2, row, "", style) #C
    write_to_sheet(sheet, 3, row, "", "thick_right") #D
    write_to_sheet(sheet, 4, row, "", centered) #E
    write_to_sheet(sheet, 5, row, "", "thick_right") #F
    write_to_sheet(sheet, 6, row, "", style) #G
    write_to_sheet(sheet, 7, row, "", "thick_right") #H
    write_to_sheet(sheet, 8, row, "", centered) #I
    write_to_sheet(sheet, 9, row, "", "thick_right") #J
    write_to_sheet(sheet, 10, row, "", centered) #K
    write_to_sheet(sheet, 11, row, "", centered) #L last updated
    write_to_sheet(sheet, 12, row, "", centered) #M
    write_to_sheet(sheet, 13, row, "", centered) #N
    write_to_sheet(sheet, 14, row, "", centered) #O
    write_to_sheet(sheet, 15, row, (data.name, data.id), centered) #P assessor link
    if conflict_message:
        write_to_sheet(sheet, 16, row, conflict_message, style) #Q
    sheet.row_dimensions[row].height = sheet.row_dimensions[row - 1].height
    return row + 1


ROW_FIELDS = ("name", "address", "coords", "conflict", "id")
Row = namedtuple("Row", ROW_FIELDS, defaults=("",) * len(ROW_FIELDS))

def remove_template_sheets(workbook):
    """ cleans up template sheets """
    for name in ("Sheet1",):
        workbook.remove(workbook[name])


def print_workbook(t_id):
    """  ready! to print """
    workbook = open_workbook("newtemplate.xlsx")
    add_default_styles(workbook)

    youngest_change = datetime(2000, 1, 1)

    for street, residences in get_territory_docs(t_id):
        residences.sort(key=key_residence)
        row = 3
        sheet = make_street_sheet(workbook, t_id, street, residences)
        printed_assessor_ids = []

        for resident in residences:
            assessor_id = resident.get("assessorAccountNumber")
            if assessor_id in printed_assessor_ids:
                continue
            printed_assessor_ids.append(assessor_id)
            assessor_link = "https://www.assessor.tulsacounty.org/Property/Info?accountNo={}".format(assessor_id)
            addr = get_gather_address(resident["address"])
            coords = [str(c) for c in resident["coords"]]

            conflict = resident.get("houseNumConflict")
            if conflict:
                addr = "*" + addr

            if youngest_change < resident.get("lastUpdate"):
                youngest_change = resident.get("lastUpdate")

            row = write_row(sheet, row, Row(assessor_link, addr, coords, conflict, assessor_id))
        sheet.row_dimensions[row + 1].height = 28
        sheet.row_dimensions[row + 2].height = 28

    remove_template_sheets(workbook)
    workbook.save(filename="{}.xlsx".format(t_id))


if __name__ == "__main__" and len(sys.argv) > 1:
    print_workbook(sys.argv[1])
